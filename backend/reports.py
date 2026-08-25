"""Report file generation — turns incident data into a real downloadable file (PDF / XLSX /
CSV), not just text to copy-paste. Used by the agent's export_report tool; the file is
served for download (web) or attached to the message (Discord)."""
import csv
import os
import time

from .config import PROJECT_ROOT
from .database import engine_lock, db_conn, _event_row_to_dict

REPORTS_DIR = os.path.join(PROJECT_ROOT, "reports")

_STATUS_LABEL = {
    "PENDING": "Menunggu Review", "CONFIRMED": "Terkonfirmasi",
    "DISMISSED": "Ditolak", "DELETED": "Dihapus",
}


def _gather(since_hours, zone):
    try:
        since_hours = float(since_hours or 24)
    except (TypeError, ValueError):
        since_hours = 24
    cutoff = time.time() * 1000 - since_hours * 3_600_000
    query = "SELECT * FROM events WHERE ts_ms >= ?"
    params = [cutoff]
    if zone:
        z = f"%{str(zone).lower()}%"
        query += " AND (LOWER(zone) LIKE ? OR LOWER(stream_id) LIKE ?)"
        params += [z, z]
    query += " ORDER BY ts_ms DESC"
    with engine_lock:
        rows = db_conn.execute(query, params).fetchall()
    events = [_event_row_to_dict(r) for r in rows]

    by_status, by_class, by_zone = {}, {}, {}
    emergency = 0
    unresolved = 0
    for e in events:
        by_status[e["status"]] = by_status.get(e["status"], 0) + 1
        by_class[e["class"]] = by_class.get(e["class"], 0) + 1
        by_zone[e["zone"]] = by_zone.get(e["zone"], 0) + 1
        if e["type"] == "EMERGENCY":
            emergency += 1
        if e["status"] == "CONFIRMED" and not e["action_taken"]:
            unresolved += 1
    summary = {
        "period_hours": since_hours,
        "zone_filter": zone or "Semua zona",
        "total": len(events),
        "by_status": by_status,
        "by_class": by_class,
        "by_zone": by_zone,
        "emergency": emergency,
        "unresolved": unresolved,
        "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    return events, summary


_COLUMNS = ["seq", "timestamp", "zone", "class", "type", "urgency", "status", "action_note"]
_HEADERS = ["#", "Waktu", "Zona", "Kelas", "Jenis", "Urgensi", "Status", "Tindakan"]


def _row(e):
    return [
        e.get("seq"), e.get("timestamp"), e.get("zone"), e.get("class"), e.get("type"),
        (e.get("urgency") or "-"), _STATUS_LABEL.get(e.get("status"), e.get("status")),
        e.get("action_note") or ("[dihapus] " + (e.get("delete_reason") or "") if e.get("deleted") else "-"),
    ]


def _to_csv(path, events, summary):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Laporan Keselamatan HSE — Smart Factory"])
        w.writerow(["Dibuat", summary["generated_at"]])
        w.writerow(["Periode (jam)", summary["period_hours"]])
        w.writerow(["Zona", summary["zone_filter"]])
        w.writerow(["Total insiden", summary["total"]])
        w.writerow(["Darurat", summary["emergency"]])
        w.writerow(["Terkonfirmasi belum ditindak", summary["unresolved"]])
        w.writerow([])
        w.writerow(_HEADERS)
        for e in events:
            w.writerow(_row(e))


def _to_xlsx(path, events, summary):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Ringkasan"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")

    ws["A1"] = "Laporan Keselamatan HSE — Smart Factory"
    ws["A1"].font = Font(bold=True, size=14)
    meta = [
        ("Dibuat", summary["generated_at"]),
        ("Periode (jam)", summary["period_hours"]),
        ("Zona", summary["zone_filter"]),
        ("Total insiden", summary["total"]),
        ("Darurat", summary["emergency"]),
        ("Terkonfirmasi belum ditindak", summary["unresolved"]),
    ]
    r = 3
    for k, v in meta:
        ws.cell(r, 1, k).font = Font(bold=True)
        ws.cell(r, 2, v)
        r += 1
    r += 1
    ws.cell(r, 1, "Per Status").font = Font(bold=True)
    r += 1
    for k, v in summary["by_status"].items():
        ws.cell(r, 1, _STATUS_LABEL.get(k, k)); ws.cell(r, 2, v); r += 1
    r += 1
    ws.cell(r, 1, "Per Zona").font = Font(bold=True)
    r += 1
    for k, v in summary["by_zone"].items():
        ws.cell(r, 1, k); ws.cell(r, 2, v); r += 1
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22

    ws2 = wb.create_sheet("Insiden")
    ws2.append(_HEADERS)
    for c in range(1, len(_HEADERS) + 1):
        cell = ws2.cell(1, c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for e in events:
        ws2.append(_row(e))
    widths = [6, 12, 26, 16, 12, 10, 18, 40]
    for i, wd in enumerate(widths, 1):
        ws2.column_dimensions[chr(64 + i)].width = wd
    ws2.freeze_panes = "A2"
    wb.save(path)


def _to_pdf(path, events, summary):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("t", parent=styles["Title"], fontSize=16, spaceAfter=6)
    doc = SimpleDocTemplate(path, pagesize=landscape(A4), topMargin=15 * mm, bottomMargin=15 * mm)
    story = [
        Paragraph("Laporan Keselamatan HSE — Smart Factory", title_style),
        Paragraph(
            f"Dibuat {summary['generated_at']} · Periode {summary['period_hours']} jam · "
            f"Zona: {summary['zone_filter']}", styles["Normal"],
        ),
        Spacer(1, 8),
    ]
    summ = [
        ["Total insiden", summary["total"], "Darurat", summary["emergency"],
         "Terkonfirmasi belum ditindak", summary["unresolved"]],
    ]
    st = Table(summ, hAlign="LEFT")
    st.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EEF2F7")),
        ("FONTNAME", (0, 0), (0, 0), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, 0), "Helvetica-Bold"),
        ("FONTNAME", (4, 0), (4, 0), "Helvetica-Bold"),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("PADDING", (0, 0), (-1, -1), 5),
    ]))
    story += [st, Spacer(1, 12), Paragraph("Daftar Insiden", styles["Heading3"])]

    data = [_HEADERS] + [[str(x) if x is not None else "" for x in _row(e)] for e in events[:400]]
    tbl = Table(data, repeatRows=1, colWidths=[14, 55, 120, 75, 60, 50, 80, 160])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("PADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(tbl)
    doc.build(story)


def generate_report_file(fmt="pdf", since_hours=24, zone=""):
    """Generate a report file. Returns {status, filename, path, format} or an error dict."""
    fmt = str(fmt).lower().strip()
    if fmt in ("excel", "xls"):
        fmt = "xlsx"
    if fmt not in ("pdf", "xlsx", "csv"):
        return {"status": "error", "message": f"Format '{fmt}' tidak didukung. Pilih pdf, xlsx, atau csv."}

    events, summary = _gather(since_hours, zone)
    os.makedirs(REPORTS_DIR, exist_ok=True)
    filename = f"laporan_hse_{time.strftime('%Y%m%d_%H%M%S')}.{fmt}"
    path = os.path.join(REPORTS_DIR, filename)
    try:
        if fmt == "csv":
            _to_csv(path, events, summary)
        elif fmt == "xlsx":
            _to_xlsx(path, events, summary)
        else:
            _to_pdf(path, events, summary)
    except Exception as e:
        return {"status": "error", "message": f"Gagal membuat laporan: {e}"}

    return {
        "status": "success",
        "filename": filename,
        "path": path,
        "format": fmt,
        "total": summary["total"],
        "period_hours": summary["period_hours"],
    }
